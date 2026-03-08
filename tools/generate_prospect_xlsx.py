"""Generate a structured XLSX file from prospect research data.

Reads a JSON file with prospect data and outputs a formatted Excel workbook
with a main prospects sheet and a summary sheet.

Usage:
    python tools/generate_prospect_xlsx.py --input .tmp/prospects.json --output .tmp/prospect_research.xlsx
"""

import argparse
import json
import logging
import os
import sys
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

PROSPECT_COLUMNS = [
    ("Brand Name", 25),
    ("Website URL", 35),
    ("Category/Vertical", 22),
    ("HQ Location", 20),
    ("Business Summary", 50),
    ("Ecommerce Platform", 18),
    ("Size Indicators", 35),
    ("Key Contact 1 (Name/Role)", 30),
    ("Key Contact 1 LinkedIn", 35),
    ("Key Contact 2 (Name/Role)", 30),
    ("Key Contact 2 LinkedIn", 35),
    ("Current Digital Agency", 22),
    ("Channel Opportunity Notes", 45),
    ("Confidence Tier", 15),
]

TIER_ORDER = {"High": 0, "Medium": 1, "Low": 2}

TIER_FILLS = {
    "High": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Low": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def validate_prospect(prospect: dict, index: int) -> list[str]:
    warnings = []
    if not prospect.get("brand_name"):
        warnings.append(f"Prospect {index}: missing brand_name")
    if not prospect.get("website_url"):
        warnings.append(f"Prospect {index}: missing website_url")
    if prospect.get("confidence_tier") not in TIER_ORDER:
        warnings.append(
            f"Prospect {index} ({prospect.get('brand_name', '?')}): "
            f"invalid confidence_tier '{prospect.get('confidence_tier')}', defaulting to Low"
        )
    return warnings


def sort_prospects(prospects: list[dict]) -> list[dict]:
    return sorted(
        prospects,
        key=lambda p: (
            TIER_ORDER.get(p.get("confidence_tier", "Low"), 2),
            (p.get("brand_name") or "").lower(),
        ),
    )


def add_hyperlink(ws, row: int, col: int, url: str, display: str | None = None):
    cell = ws.cell(row=row, column=col)
    if url and url.startswith(("http://", "https://")):
        cell.hyperlink = url
        cell.value = display or url
        cell.font = LINK_FONT
    else:
        cell.value = url or ""


def create_prospects_sheet(wb: Workbook, prospects: list[dict]):
    ws = wb.active
    ws.title = "Prospects"

    # Write headers
    for col_idx, (header, width) in enumerate(PROSPECT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write data
    sorted_prospects = sort_prospects(prospects)
    for row_idx, prospect in enumerate(sorted_prospects, 2):
        tier = prospect.get("confidence_tier", "Low")
        if tier not in TIER_ORDER:
            tier = "Low"
        tier_fill = TIER_FILLS.get(tier)

        row_data = [
            prospect.get("brand_name", ""),
            prospect.get("website_url", ""),
            prospect.get("category", ""),
            prospect.get("hq_location", ""),
            prospect.get("business_summary", ""),
            prospect.get("ecommerce_platform", ""),
            prospect.get("size_indicators", ""),
            prospect.get("key_contact_1_name_role", ""),
            prospect.get("key_contact_1_linkedin", ""),
            prospect.get("key_contact_2_name_role", ""),
            prospect.get("key_contact_2_linkedin", ""),
            prospect.get("current_agency", ""),
            prospect.get("channel_opportunity_notes", ""),
            tier,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

        # Add hyperlinks for URL columns
        add_hyperlink(ws, row_idx, 2, prospect.get("website_url", ""))
        add_hyperlink(
            ws, row_idx, 9,
            prospect.get("key_contact_1_linkedin", ""),
            prospect.get("key_contact_1_linkedin", ""),
        )
        add_hyperlink(
            ws, row_idx, 11,
            prospect.get("key_contact_2_linkedin", ""),
            prospect.get("key_contact_2_linkedin", ""),
        )

        # Apply tier colour to the confidence column
        if tier_fill:
            ws.cell(row=row_idx, column=14).fill = tier_fill

    # Set row height for data rows
    for row_idx in range(2, len(sorted_prospects) + 2):
        ws.row_dimensions[row_idx].height = 45


def create_summary_sheet(
    wb: Workbook,
    prospects: list[dict],
    searched_verticals: list[str] | None = None,
    empty_verticals: list[str] | None = None,
):
    ws = wb.create_sheet("Summary")

    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    section_font = Font(name="Calibri", size=12, bold=True, color="1F4E79")

    # Title
    ws.cell(row=1, column=1, value="Prospect Research Summary").font = title_font
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15

    # Total count
    row = 3
    ws.cell(row=row, column=1, value="Total Prospects").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(prospects))

    # Count by confidence tier
    row = 5
    ws.cell(row=row, column=1, value="By Confidence Tier").font = section_font
    row += 1
    tier_counts = Counter(p.get("confidence_tier", "Low") for p in prospects)
    for tier in ["High", "Medium", "Low"]:
        ws.cell(row=row, column=1, value=tier)
        ws.cell(row=row, column=2, value=tier_counts.get(tier, 0))
        if tier in TIER_FILLS:
            ws.cell(row=row, column=1).fill = TIER_FILLS[tier]
        row += 1

    # Count by category
    row += 1
    ws.cell(row=row, column=1, value="By Category/Vertical").font = section_font
    row += 1
    category_counts = Counter(p.get("category", "Uncategorised") for p in prospects)
    for category, count in sorted(category_counts.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=category)
        ws.cell(row=row, column=2, value=count)
        row += 1

    # Empty verticals
    if empty_verticals:
        row += 1
        ws.cell(row=row, column=1, value="Verticals With No Results").font = section_font
        row += 1
        for vertical in empty_verticals:
            ws.cell(row=row, column=1, value=vertical)
            row += 1

    # Notes
    row += 1
    ws.cell(row=row, column=1, value="Notes").font = section_font
    row += 1
    ws.cell(
        row=row, column=1,
        value="High = Confirmed WA HQ, confirmed ecommerce, revenue indicators suggest $2M-$50M",
    )
    row += 1
    ws.cell(
        row=row, column=1,
        value="Medium = Likely WA-based with ecommerce, but revenue/spend unconfirmed",
    )
    row += 1
    ws.cell(
        row=row, column=1,
        value="Low = Possible fit, needs manual validation",
    )


def generate_xlsx(
    prospects: list[dict],
    output_path: str,
    searched_verticals: list[str] | None = None,
    empty_verticals: list[str] | None = None,
):
    wb = Workbook()
    create_prospects_sheet(wb, prospects)
    create_summary_sheet(wb, prospects, searched_verticals, empty_verticals)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    logger.info(f"XLSX saved to {output_path}")
    logger.info(f"  {len(prospects)} prospects across 2 sheets")
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description="Generate prospect research XLSX from JSON data"
    )
    parser.add_argument(
        "--input", required=True, help="Path to JSON file with prospect data"
    )
    parser.add_argument(
        "--output", required=True, help="Path for output XLSX file"
    )
    args = parser.parse_args()

    if not os.path.exists(args.input):
        logger.error(f"Input file not found: {args.input}")
        sys.exit(1)

    with open(args.input) as f:
        data = json.load(f)

    if isinstance(data, list):
        prospects = data
        searched_verticals = None
        empty_verticals = None
    elif isinstance(data, dict):
        prospects = data.get("prospects", [])
        searched_verticals = data.get("searched_verticals")
        empty_verticals = data.get("empty_verticals")
    else:
        logger.error("Input JSON must be a list of prospects or an object with a 'prospects' key")
        sys.exit(1)

    if not prospects:
        logger.error("No prospects found in input data")
        sys.exit(1)

    # Validate
    all_warnings = []
    for i, prospect in enumerate(prospects):
        all_warnings.extend(validate_prospect(prospect, i))
    if all_warnings:
        for w in all_warnings:
            logger.warning(f"  Warning: {w}")

    generate_xlsx(prospects, args.output, searched_verticals, empty_verticals)


if __name__ == "__main__":
    main()
